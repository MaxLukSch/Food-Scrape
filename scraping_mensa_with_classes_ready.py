# TUTORIAL WEB SCRAPING http://python-guide-pt-br.readthedocs.io/en/latest/scenarios/scrape/

from lxml import html
import requests
from openpyxl import load_workbook
import time

class Restaurants(object):                  
    def __init__(self, name, link):  # input variables for new instance (restaurant)
        self.name = name           
        self.link = link
    
    def get_meals(self):
        # goal of the function is to get lists which all have the length of the amount of meals
        # which are served today. there are lists for the meals, their prices, the date, the amount
        # of sides, the type of extras available. these lists are saved into an excel file. the
        # script should be run only during weekdays. since it always takes the first meals avaiable
        # and thinks they are served today. this might be fixed soon.
        
        # get the content of the page
        page = requests.get("%s" % self.link)
        tree = html.fromstring(page.content)

        # get the meals and prices for the next week
        meals = tree.xpath('//strong[@class="menu_name"]/text()')
        prices = tree.xpath('//p[@class="pull-right"]/strong/text()')

        # get the meals for today and their prices
        dates = tree.xpath('//div[@class="panel panel-default"]/div[@class="panel-heading"]/strong/text()')
        meals_minus_day_one = tree.xpath('//strong[@class="menu_name"][preceding::div[@class="panel panel-default"][contains(.,"%s")]]/text()' % dates[0])
        all_meals = len(meals)
        amount_minus_day_one = len(meals_minus_day_one)
        amount_today = all_meals - amount_minus_day_one
        
        meals_today = meals[0:amount_today]
        prices_today = prices[0:amount_today] 

        # remove the "€" from the prices
        prices_saveable = []

        for price in prices_today:
            string = price.replace("€", "")
            string2 = string.replace(" ", "")
            prices_saveable.append(string2)

        # get number of sides and extras

        # get the whole content (sides, extras and lots of unneeded code)
        sides_raw = []

        for i in range(0, len(meals_today)): 
            side = tree.xpath('//p[following::strong[@class="menu_name"][contains(.,"%s")] and preceding::strong[@class="menu_name"][contains(.,"%s")]]/text()' % (meals[i+1], meals[i]))
            sides_raw.append(side)


        # create list with number of sides

        count_sides = []

        for i in range(0, len(sides_raw)):
            if "einer Beilage" in str(sides_raw[i]):
                count_sides.append(1)

            elif "zwei Beilagen" in str(sides_raw[i]):
                count_sides.append(2)
            
            else:
                count_sides.append(0)


        # remove the words "Beilage" and "Wahl" from the sides_raw code, in order to look for the extras, which starte with capital letters

        sides_minus_BaW = []

        import re

        for i in range(0, len(sides_raw)):
            temp = ""
            temp = re.sub(r'Beilage', ' ' ,re.sub(r'Wahl', r'.  ', str(sides_raw[i])))
            sides_minus_BaW.append(temp)

        # identify the extras, extract their first (capital) letter

        import string
        alpha = []
                                                                        
        for i in range(26):
            alpha.extend(string.ascii_uppercase[i:i+1])

        extras_first = {}


        for y in range(0, len(sides_minus_BaW)):
            for i in range(0, len(alpha)):
                if alpha[i] in str(sides_minus_BaW[y]):
                    extras_first[y] = alpha[i]
                    break
                else:
                    extras_first[y] = "None"

        # which meal has an extra (get its number in the meals from today

        numbers = []
        for i in range(0, len(extras_first)):
            if extras_first[i] != "None":
                numbers.append(i)

        # get the names of the extras (without the first letter)
        # lamdba function useful for a loop inside a loop

        extras = []

        for y in range(0, len(sides_minus_BaW)):
            temporary = ""
            for i in range(0, len(alpha)):
                if alpha[i] in str(sides_minus_BaW[y]):
                    funct = lambda x: sides_minus_BaW[y][sides_minus_BaW[y].find(x)+1:].split()[0]
                    temporary = funct(extras_first[y])
                    extras.append(temporary)

        # turn the the dict into a list
        extras_first_list = list(extras_first.values())

        # create a dict, which has the the names of the extras (without the first letter)
        # first create new dict with just the meals that have no extra and then add at the correct number the extras

        # this is called dict comprehension. remove certain keys from a dict
        extras_just_nones = {k:v for k,v in extras_first.items() if v is "None"}

        counter = 0
        for i in numbers:
            extras_just_nones[i] = extras[counter]
            counter += 1

        #sort this new dict according 1,2,3 etc. this automatically turns it into a list

        extras_just_nones_s = sorted(extras_just_nones.items())

        # merge the lists with the first letter and the rest of the extras names

        extras_whole = []
        
        counter2 = 0
        for i in range(0, len(extras_first_list)):
            extras_whole.append(extras_first_list[i] + extras_just_nones_s[counter2][1])
            counter2 += 1

        # "extras_whole" is a list with the length of the other lists like meals , dates , price, sides etc.
        
        

        # save the date, meals, prices, sides, extras 

        # open the excelfile and choose the right sheet - which beforehand has to be created
        wb = load_workbook('mensa.xlsx')
        ws = wb["%s" % self.name] 

        # check to which row the sheet is already filled
        m = ws.max_row
        p = ws.max_row
        d = ws.max_row
        s = ws.max_row
        e = ws.max_row
        
        # get todays date and save it
        date = time.strftime("%x")
        
        for meal in meals_today:
            ws.cell(row= d, column=1).value = date
            d += 1

        # save the meals
        for meal in meals_today:
            ws.cell(row= m, column=2).value = meal
            m += 1

        # save the prices
        for price in prices_saveable:
            ws.cell(row= p, column=3).value = price
            p += 1

        # save the sides
        for side in count_sides:
            ws.cell(row= s, column=4).value = side
            s += 1

        # save the extras
        for extra in extras_whole:
            ws.cell(row= e, column=5).value = extra
            e += 1

        # save the file
        wb.save("mensa.xlsx")

     
darwins = Restaurants("darwins", "https://www.studentenwerkfrankfurt.de/essen-trinken/speiseplaene/cafeteria-darwins/")
pixgaumen = Restaurants("pixgaumen", "https://www.studentenwerkfrankfurt.de/essen-trinken/speiseplaene/mensa-pi-x-gaumen/")

darwins.get_meals()
pixgaumen.get_meals()






