# -*- coding: utf-8 -*-
# the first line is necessary for the use of german umlaute

import requests #V 2.18.3
from bs4 import BeautifulSoup #V 4.6.0
from openpyxl import load_workbook #V 2.4.8
import re
import string
import time

class Restaurants(object):                  
    def __init__(self, name, link):  # input variables for new instance (restaurants)
        self.name = name           
        self.link = link
    
    def get_meals(self):
        # this function aims at retrieving information on he meals served
        # in the restaurants of the Studentwerk Frankfurt. there are lists
        # retrieved for the meals, their prices, the amount of sides dishes
        # and extra dishes. the script has to be run on every operating day
        # of the restaurants.
        
   
        # grab the page
        result = requests.get("%s" % self.link)

        # store it in an easily accesible variable
        c = result.content

        # parse with bs - set the parser to make sure it works everywhere the same way
        soup = BeautifulSoup(c, "html.parser")

        # remove all the "span" elements of the code - theyve been causing trouble
        for span in soup("span"):
            span.decompose()


        # extract information on meals and prices
        meals = soup.find_all("strong", "menu_name")
        prices = soup.find_all("p", "pull-right")

        # get the meals list. extract the string via .string

        list_meals = []

        for i in range(0, len(meals)):
            temp = meals[i]
            list_meals.append(temp.string)

        # get the prices list. therefore extract the children of prices list which contains it first.

        list_prices = []

        for i in range(0, len(prices)):
            temp = prices[i].contents[1]
            temp2 = temp.string
            list_prices.append(temp2)

        # retrive information on side dishes and extras in "rest"

        rest = []

        for i in range(0, len(meals)):
            temp = [element.text for element in meals[i].find_next_siblings("p")]
            rest.append(temp)

        # delete words shorter than two letters and backslashes

        reduced = []

        for i in range(0, len(rest)):
            temp = ""    
            temp = re.sub(r'\b\w{1,2}\b', '', str(rest[i]))
            reduced.append(temp)


        reduced2 = []

        for i in range(0, len(reduced)):
            temp = ""    
            temp = re.sub(r'\\', '', str(reduced[i])) 
            reduced2.append(temp)

        # get the number of side dishes

        list_sides = []

        for i in range(0, len(reduced)):
            if "einer Beilage" in str(reduced[i]):
                list_sides.append(1)

            elif "zwei Beilagen" in str(reduced[i]):
                list_sides.append(2)
            
            else:
                list_sides.append(0)
    
        # delete info on side dishes

        reduced3 = []

        for i in range(0, len(reduced2)):
            temp = ""    
            temp = re.sub(r'tmit einer Beilage nach Wahl', ' ' ,re.sub(r'tund einer Beilage nach Wahl', r'.  ', str(reduced2[i])))
            reduced3.append(temp)

        reduced4 = []

        for i in range(0, len(reduced3)):
            temp = ""    
            temp = re.sub(r'tmit zwei Beilagen nach Wahl', ' ' ,re.sub(r'tund zwei Beilagen nach Wahl', r'.  ', str(reduced3[i])))
            reduced4.append(temp)

        # delete the trash in the beginning " ["t "

        reduced5 = []

        for i in range(0, len(reduced4)):
            temp = ""
            temp = str(reduced4[i])
            reduced5.append(temp[3:])

        # delete the trash at the end. use "alpha" in order to delete unnecessary letters

        alpha = []
                                                                        
        for i in range(26):
            alpha.extend(string.ascii_lowercase[i:i+1])

        for i in range(26):
            alpha.extend(string.ascii_uppercase[i:i+1])

        alpha.append(" ")
        alpha.append("-")
        alpha.append("ß")
        alpha.append("ä")
        alpha.append("ö")
        alpha.append("ü")

        # gather the information in "extras_list"

        list_extras_w = []

        for i in range(0, len(reduced5)):
            temp = ""
            for letter in reduced5[i]:
                if letter in alpha:
                    temp += letter
            list_extras_w.append(temp)

        # now delete all the unnecessary whitespace at the end of the extras
        # string.strip () does this conveniently

        list_extras = []

        for i in range(0, len(list_extras_w)):
            temp = str(list_extras_w[i])
            temp2 = temp.strip()
            list_extras.append(temp2)

        # which meals are served today? necessary because this changes on the day
        # they are served. therefore you cannot save information on the meals that
        # will be served tomorrow. due to the changes tomorrow youre data will be
        # incomplete

        # in the HTML every day starts with a "table". extract the first "table" and count the meals

        am = soup.table

        col = am.find_all("strong", "menu_name")

        amount_today = len(col)

        # finalize lists

        meals_today = list_meals[0:amount_today]

        prices_today = list_prices[0:amount_today]

        sides_today = list_sides[0:amount_today]

        extras_today = list_extras[0:amount_today]



        # save the lists into an excelfile

        # open the excelfile and choose a sheet - which beforehand has to be created
        wb = load_workbook('mensa2/meals.xlsx')
        ws = wb["%s" % self.name] 

        # check to which row the sheet is already filled
        
        d = ws.max_row
        m = ws.max_row
        p = ws.max_row
        s = ws.max_row
        e = ws.max_row

        # get todays date and save it
        date = time.strftime("%x")
        
        for meal in meals_today:
            ws.cell(row= d, column=1).value = date
            d += 1

        # save the meals
        for meal in meals_today:
            ws.cell(row= m, column=3).value = meal
            m += 1

        # save the prices
        for price in prices_today:
            ws.cell(row= p, column=4).value = price
            p += 1

        # save the sides
        for side in sides_today:
            ws.cell(row= s, column=5).value = side
            s += 1

        # save the extras
        for extra in extras_today:
            ws.cell(row= e, column=6).value = extra
            e += 1

        # save the file
        wb.save("mensa2/meals.xlsx")

     
# usage, for example:

darwins = Restaurants("darwins", "https://www.studentenwerkfrankfurt.de/essen-trinken/speiseplaene/cafeteria-darwins/")
darwins.get_meals()

pixgaumen = Restaurants("pixgaumen", "https://www.studentenwerkfrankfurt.de/essen-trinken/speiseplaene/mensa-pi-x-gaumen/")
pixgaumen.get_meals()



