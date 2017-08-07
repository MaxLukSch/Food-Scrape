# TUTORIAL WEB SCRAPING http://python-guide-pt-br.readthedocs.io/en/latest/scenarios/scrape/

from lxml import html #V 3.8.0
import requests #V 2.18.2
from openpyxl import load_workbook #V 2.4.8
import time

class Restaurants(object):                  
    def __init__(self, name, link):  # input variables for new instance (restaurant)
        self.name = name           
        self.link = link
    
    def get_meals(self):
        # goal of the function is to get lists which all have the length of the amount of meals
        # which are served today. there are lists for the meals, their prices and the dates.
        # these lists are saved into an excel file. the script should be run only during weekdays.
        # since it always takes the first meals avaiable
        # and thinks they are served today. this might be fixed soon.
        
        # get the content of the page
        page = requests.get("%s" % self.link)
        tree = html.fromstring(page.content)

        # get the meals and prices for the next week
        meals = tree.xpath('//strong[@class="menu_name"]/text()')
        prices = tree.xpath('//p[@class="pull-right"]/strong/text()')

      
        # get the meals for today and their prices
        dates = tree.xpath('//div[@class="panel panel-default"]/div[@class="panel-heading"]/strong/text()')
        meals_minus_day_one = tree.xpath('//strong[@class="menu_name"][preceding::div[@class="panel panel-default"][contains(.,"%s")]]/text()' % dates[0])
        all_meals = len(meals)
        amount_minus_day_one = len(meals_minus_day_one)
        amount_today = all_meals - amount_minus_day_one
        
        meals_today = meals[0:amount_today]
        prices_today = prices[0:amount_today] 

        # remove the "€" from the prices
        prices_saveable = []

        for price in prices_today:
            string = price.replace("€", "")
            string2 = string.replace(" ", "")
            prices_saveable.append(string2)


        # save the date, meals and prices 

        # open the excelfile and choose the right sheet - which beforehand has to be created
        wb = load_workbook('mensa/meals.xlsx')
        ws = wb["%s" % self.name] 

        # check to which row the sheet is already filled
        m = ws.max_row
        p = ws.max_row
        d = ws.max_row
              
        # get todays date and save it
        date = time.strftime("%x")
        
        for meal in meals_today:
            ws.cell(row= d, column=1).value = date
            d += 1

        # save the meals
        for meal in meals_today:
            ws.cell(row= m, column=2).value = meal
            m += 1

        # save the prices
        for price in prices_saveable:
            ws.cell(row= p, column=3).value = price
            p += 1

        # save the file
        wb.save("mensa/meals.xlsx")

     

# the scripts works for all the restaurants of the Studenwerk Frankfurt
# for example:

darwins = Restaurants("darwins", "https://www.studentenwerkfrankfurt.de/essen-trinken/speiseplaene/cafeteria-darwins/")
darwins.get_meals()





